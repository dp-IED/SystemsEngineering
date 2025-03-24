# test_integration_pipeline.py
import time
from azure.storage.blob import BlobServiceClient
from openpyxl import load_workbook
import os

BLOB_CONNECTION_STRING = "DefaultEndpointsProtocol=https;AccountName=systemsteam17storage;AccountKey=Wr7IgB+c6ghclYn9rcwRgNpYv16cVKe/0hUWtS1GD/wCcosZcVfFQ0UshCwir6QAykXqfFkcpBVN+AStgDyYYQ==;EndpointSuffix=core.windows.net"
DOCUMENT_NAME = "Chanel UK Billed.xlsx"
CSV_CONTAINER = "csv-conversion"
SUMMARY_CONTAINER = "summary"

def upload_test_doc():
    blob_service = BlobServiceClient.from_connection_string(BLOB_CONNECTION_STRING)
    container = blob_service.get_container_client("subcontractor-documents")
    with open(DOCUMENT_NAME, "rb") as data:
        container.upload_blob(DOCUMENT_NAME, data, overwrite=True)
    print("Uploaded test document.")

def wait_for_file(container_name, file_name, timeout=60):
    blob_service = BlobServiceClient.from_connection_string(BLOB_CONNECTION_STRING)
    container = blob_service.get_container_client(container_name)
    for _ in range(timeout):
        if any(blob.name == file_name for blob in container.list_blobs()):
            print(f"{file_name} is available in {container_name}")
            return True
        time.sleep(2)
    raise TimeoutError(f"{file_name} not found in {container_name} after {timeout} seconds.")

def validate_excel():
    blob_service = BlobServiceClient.from_connection_string(BLOB_CONNECTION_STRING)
    container = blob_service.get_container_client(SUMMARY_CONTAINER)
    blob_name = "FormattedAnnualBudget.xlsx"
    blob_client = container.get_blob_client(blob_name)

    # Download to disk
    with open(blob_name, "wb") as f:
        f.write(blob_client.download_blob().readall())
    print("Downloaded Excel file.")

    # ✅ Load and inspect
    try:
        wb = load_workbook(blob_name)
        assert len(wb.sheetnames) > 0, "Excel file has no sheets"

        expected_sheets = ["F&B Monthly", "FSH&EW Monthly", "W&FJ Monthly"]
        for sheet in expected_sheets:
            assert sheet in wb.sheetnames, f"Expected sheet '{sheet}' not found in Excel file"

        print(f"Excel file contains sheets: {wb.sheetnames}")
        print("Excel validation passed ✅")
    except Exception as e:
        raise AssertionError(f"Excel validation failed ❌: {e}")
    finally:
        os.remove(blob_name)


if __name__ == "__main__":
    upload_test_doc()
    wait_for_file(CSV_CONTAINER, "billed_report.csv")  # or whatever it gets renamed to
    wait_for_file(SUMMARY_CONTAINER, "FormattedAnnualBudget.xlsx")
    validate_excel()
    print("Integration test passed ✅")
