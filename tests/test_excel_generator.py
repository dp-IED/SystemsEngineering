import unittest
import pandas as pd
import tempfile
import os
from openpyxl import load_workbook
from AdxIngestFunction.excel_generator import generate_excel_report

class TestExcelGenerator(unittest.TestCase):

    def setUp(self):
        """Create a temp Excel file path for testing."""
        self.tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name

    def tearDown(self):
        """Clean up the temporary file after each test."""
        if os.path.exists(self.tmp_file):
            os.remove(self.tmp_file)

    def test_generate_excel_report(self):
        """Test if the Excel report is properly generated and formatted."""
        # Create test DataFrame with multiple divisions
        df = pd.DataFrame({
            'Division': ['F&B', 'FSH&EW', 'W&FJ', 'F&B'],
            'Channel': ['Online', 'Retail', 'Wholesale', 'Retail'],
            'TotalInvoiceVal': [1000, 2000, 1500, 1800],
            'POValueRemaining': [500, 700, 300, 600],
            'POCloseDownDate': ['2024-03-10', '2024-04-15', '2024-05-20', '2024-06-05']
        })

        # Save DataFrame as CSV (Simulate input data)
        temp_csv = tempfile.NamedTemporaryFile(delete=False, suffix=".csv").name
        df.to_csv(temp_csv, index=False)

        try:
            # Run function to generate the Excel file
            generate_excel_report(self.tmp_file)

            # Load generated Excel file
            wb = load_workbook(self.tmp_file)

            # Check if sheets were created correctly
            expected_sheets = ['F&B', 'FSH&EW', 'W&FJ']
            for sheet in expected_sheets:
                self.assertIn(sheet, wb.sheetnames, f"Sheet {sheet} not found in workbook")

            # Check column deletion logic by ensuring 'Division' column is removed
            ws = wb['F&B']
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            self.assertNotIn("Division", headers, "Column 'Division' should have been deleted")

            # Check if header row formatting was applied
            for cell in ws[1]:  
                self.assertIsNotNone(cell.fill, "Header fill is missing")
                self.assertIsNotNone(cell.font, "Header font is missing")

            # Check merged cell logic (Assume it's based on some aggregation)
            merged_ranges = [str(range_obj) for range_obj in ws.merged_cells.ranges]
            self.assertGreater(len(merged_ranges), 0, "No merged cells found, expected at least one")

        finally:
            os.remove(temp_csv)  # Cleanup CSV file

if __name__ == '__main__':
    unittest.main()
