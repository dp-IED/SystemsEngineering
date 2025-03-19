import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from azure.kusto.data import KustoClient, KustoConnectionStringBuilder
from azure.kusto.data.helpers import dataframe_from_result_table
import io
import os
import time

def save_dataframe_to_excel(df, filename, divisions):
    """
    Saves the DataFrame to an Excel file, splitting it into sheets by divisions.
    """
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for division in divisions:
            df_division = df[df['Division'] == division]
            sheet_name = str(division).replace(':', '').replace('\\', '').replace('/', '') \
                                      .replace('?', '').replace('*', '')[:31]
            if sheet_name.strip() == '':
                sheet_name = 'Other'
            df_division.to_excel(writer, sheet_name=sheet_name, index=False)

def delete_column_from_sheets(workbook, column_title):
    """
    Deletes a column from all sheets in the workbook based on the column title.
    """
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        column = None
        for cell in ws[1]:
            if cell.value == column_title:
                column = cell.column
                break
        if column:
            ws.delete_cols(column)

def format_workbook(filename):
    """
    Formats the Excel workbook by adjusting column widths.
    """
    wb = load_workbook(filename)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
    wb.save(filename)

# Define your database and table
CLUSTER = "https://chanelmediacluster.uksouth.kusto.windows.net"
DATABASE = "financial-database-1"

# Authentication
KCSB = KustoConnectionStringBuilder.with_az_cli_authentication(CLUSTER)
client = KustoClient(KCSB)

# Define queries
QUERY_1 = """
unbilled_report
// add column of months
| extend Month = datetime_part("month", todatetime(BuyMonth))
| extend Month = case(
Month == 1, "January",
Month == 2, "February",
Month == 3, "March",
Month == 4, "April",
Month == 5, "May",
Month == 6, "June",
Month == 7, "July",
Month == 8, "August",
Month == 9, "September",
Month == 10, "October",
Month == 11, "November",
Month == 12, "December",
"Unknown"
)
// add column of media names from PO tracker
| extend Channel = case(
MediaName == "SEARCH&SOC", "Social & PPC",
MediaName == "DISPLAY", "Display & VOD",
MediaName == "PRESS", "Print",
MediaName == "CINEMA", "Cinema",
MediaName == "TELEVISION", "TV",
MediaName == "FEES", "Retainer Fee",
MediaName == "POSTER", "OOH",
MediaName == "INT'L", "Retainer Fee",
MediaName)
// normalise product names
| extend NormalisedProductName = case(
ProductName == "EYEWEAR", "Eyewear",
ProductName contains "FASHION" or ProductName contains "MDA", "Fashion",
ProductName == "JEWELLERY HJ", "High Jewellery",
ProductName contains "WATCHES", "Watches",
ProductName == "BLEU", "Bleu",
ProductName == "UK", "PPC",
ProductName contains "MAKE UP" or ProductName contains "LES BEIGE", "Make Up",
ProductName contains "SKINCARE", "Skincare",
ProductName contains "CHANCE", "Chance",
ProductName contains "COCO MELLE", "Coco Melle",
ProductName contains "NO 5", "No 5",
ProductName)
// create market column based on LocalCurrencyCode
| extend Market = case(
CampaignName has "UK", "UK",
CampaignName has "IRE", "IRE",
"")
// clean up campaign name - remove _ and CHANEL, remove channel name
| extend Campaign = replace_regex(CampaignName, @"_", " ")
| extend Campaign = replace_regex(Campaign, @"\b(CHANEL)\b", "")
| extend Campaign = replace_regex(Campaign, @"-", " ")
| extend Campaign = replace_regex(Campaign, @"\b(UK|IRE|SOCIAL|SOCIA|SO|SOC|DISPLAY|D|DSP|DISP|FEE|FEES)\b", "")
// create year column, and create MainName column without year and suffix
| extend Year = extract(@"(20\d{2}|2[0-9])\b", 0, Campaign)
| extend Year = iff(strlen(Year) == 2, strcat("20", Year), Year)
// remove unnecessary characters
| extend Campaign = trim(" ", replace_regex(Campaign, @"(20\d{2}|2[0-9])\b|FB|FR|MU|WFJ|\bWA\b|FSH|BEA|VIDEO|\.", ""))
// remove spaces
| extend Campaign = trim(" ", replace_regex(Campaign, @" ", ""))
// normalise campaign names
| extend Campaign = case(
Campaign contains "BLEUH1", "Bleu H1",
Campaign contains "BLEUH2", "Bleu H2",
Campaign contains "LESBEIGEHERO" or Campaign contains "LESBEIGESHERO", "Les Beiges HERO",
Campaign contains "COCOCRUSHH1", "Coco Crush H1",
Campaign contains "J12H1", "J12 H1",
Campaign contains "N1H2", "N1 H2",
Campaign contains "EYEWEAR", "Eyewear",
Campaign contains "NO5LEAU", "No.5 Leau",
Campaign contains "C50/07", "C50/07",
Campaign contains "NO5LEAUDROP", "No.5 Leau Drop",
Campaign contains "BLEU", "Bleu",
Campaign contains "LESBEIGES" or Campaign contains "LESBEIGE", "Les Beiges",
Campaign contains "HIGHJEWELLERY", "High Jewellery",
Campaign contains "COCOMELLE", "Coco Melle",
ProductName == "MDA", "MDA",
Campaign)
| extend Division = case(
Channel == "Retainer Fee" and Campaign contains "F&B", "F&B",
Channel == "Retainer Fee" and Campaign contains "FASHION", "FSH&EW",
Channel == "Retainer Fee" and Campaign contains "EYEWEAR", "FSH&EW",
Channel == "Retainer Fee" and Campaign contains "JEWELLERY", "W&FJ",
Channel == "Retainer Fee" and Campaign contains "WATCHES", "W&FJ",
Division == "Paid Search", "F&B",
Channel == "Retainer Fee", Division, // Keep the original division if no match
Division // Default to the original division if Channel is not "Retainer Fee"
)
| summarize
NetBillable = sum(Payable),
AgencyCommission = sum(AgencyCommission),
LevyASBOF = sum(LevyBillable),
ProductCode = take_any(ProductCode),
NormalisedProductName = take_any(NormalisedProductName),
PO_Number = take_any(PO),
Market = take_any(Market),
Total_Invoice_Val = sum(UnbilledClientCost)
by Campaign, Division, Channel, Month
| sort by Division asc, Campaign asc
"""

QUERY_2 = """
budget_tracker
...
| project Campaign, PlannedSpend, ReservedBudget, TotalBudget
"""

QUERY_2 = """
budget_tracker
| extend GrandTotalReserve = ['GRAND TOTAL inc reserve'],
PlannedSpendData = ['Planned Spend latest plan'],
Reserve = ['Reserve'] // Include the reserve amount
| extend TotalAvailableBudget = GrandTotalReserve // Budget already includes reserve
| extend Campaign = tolower(trim(" ", Campaign))
| join kind=leftouter (
unbilled_report
| extend Campaign = replace_regex(CampaignName, @"_", " ")
| extend Campaign = replace_regex(Campaign, @"\b(CHANEL)\b", "")
| extend Campaign = replace_regex(Campaign, @"-", " ")
| extend Campaign = replace_regex(Campaign, @"\b(UK|IRE|SOCIAL|DISPLAY|FEE|DSP|DISP)\b", "")
| extend Campaign = trim(" ", replace_regex(Campaign, @"(20\d{2})\b|FB|FR|MU|WFJ|WA|FSH|BEA|VIDEO|\.", ""))
| extend Campaign = trim(" ", replace_regex(Campaign, @"\s+", " "))
| extend Campaign = tolower(Campaign)
| extend Campaign = case(
Campaign contains "bleu", "Bleu",
Campaign contains "no.5", "No. 5",
Campaign contains "les beige", "Make Up",
Campaign contains "coco melle", "Coco Melle",
Campaign contains "chance", "Chance",
Campaign contains "skincare", "Skincare",
Campaign contains "travel", "Travel Retail",
Campaign contains "watches", "Watches",
Campaign contains "jewellery", "Fine Jewellery",
Campaign contains "high jewellery", "High Jewellery",
Campaign contains "fashion", "Fashion",
Campaign contains "eyewear", "Eyewear",
Campaign contains "ppc", "PPC",
"Other")
| project Campaign, UnbilledCurrency
) on Campaign
| summarize
TotalUnbilled = sum(UnbilledCurrency),
TotalBudget = sum(TotalAvailableBudget),
PlannedSpend = sum(PlannedSpendData),
ReservedBudget = sum(Reserve)
by Campaign
| project Campaign, PlannedSpend, ReservedBudget, TotalBudget
"""

QUERY_3 = "annual_budget_sheet"

QUERY_4 ="""unbilled_report
| extend Month = datetime_part("month", todatetime(BuyMonth))
| extend MonthName = case(
Month == 1, "January",
Month == 2, "February",
Month == 3, "March",
Month == 4, "April",
Month == 5, "May",
Month == 6, "June",
Month == 7, "July",
Month == 8, "August",
Month == 9, "September",
Month == 10, "October",
Month == 11, "November",
Month == 12, "December",
"Unknown"
)
| project-away Month
| extend Campaign = tostring(CampaignName)
| extend Channel = case(
MediaName == "SEARCH&SOC", "Social & PPC",
MediaName == "DISPLAY", "Display & VOD",
MediaName == "PRESS", "Print",
MediaName == "CINEMA", "Cinema",
MediaName == "TELEVISION", "TV",
MediaName == "FEES", "Retainer Fee",
MediaName == "POSTER", "OOH",
MediaName == "INT'L", "Retainer Fee",
MediaName)
| extend Market = case(
CampaignName has "UK", "UK",
CampaignName has "IRE", "IRE",
""
)
| extend Campaign = replace_regex(Campaign, @"\b(_|UK|IRE|SOCIAL|DISPLAY|D|DSP|DISP|FEE|FEES)\b", "") // Clean up campaign names
| summarize
NetBillable = sum(Payable),
AgencyCommission = sum(AgencyCommission),
LevyASBOF = sum(LevyBillable),
Total_Invoice_Val = sum(UnbilledClientCost) // Assuming UnbilledClientCost is the correct field name for total invoice value
by MonthName, Campaign, Channel, Market
| order by MonthName, Campaign, Market
"""

# Execute queries
response_1 = client.execute(DATABASE, QUERY_1)
response_2 = client.execute(DATABASE, QUERY_2)
response_3 = client.execute(DATABASE, QUERY_3)
response_4 = client.execute(DATABASE, QUERY_4)

# Convert responses to DataFrames
cleaned_unbilled = dataframe_from_result_table(response_1.primary_results[0])
cleaned_budget_tracker = dataframe_from_result_table(response_2.primary_results[0])
annual_budget = dataframe_from_result_table(response_3.primary_results[0])
monthly_table = dataframe_from_result_table(response_4.primary_results[0])

# Group and summarize data
annual_summary = cleaned_unbilled.groupby(["Campaign", "Channel", "Division", "Month"], as_index=False).agg({
    "Division": "first",
    "NetBillable": "sum",
    "AgencyCommission": "sum",
    "LevyASBOF": "sum",
    "ProductCode": "first",
    "NormalisedProductName": "first",
    "PO_Number": lambda x: ", ".join(sorted(set(x.dropna()))),
    "Market": lambda x: ", ".join(sorted(set(x.dropna()))),
    "Total_Invoice_Val": "sum"
})

# Append total rows for each campaign
overall_totals = annual_summary.groupby(['Campaign', 'Division']).agg({
    'NetBillable': 'sum',
    'AgencyCommission': 'sum',
    'LevyASBOF': 'sum',
    'Total_Invoice_Val': 'sum',
    'ProductCode': 'first',
    'NormalisedProductName': 'first',
    'PO_Number': 'first',
    'Market': 'first'
}).reset_index()

overall_totals['Channel'] = "Total"
overall_totals['Month'] = "Total"

# Combine detailed and total rows
unmatched_df = pd.concat([annual_summary, overall_totals], ignore_index=True)
unmatched_df = unmatched_df.sort_values(by=['Division', 'Campaign', 'Month']).reset_index(drop=True)

# Normalize campaign names
unmatched_df["NormalisedProductName"] = unmatched_df["NormalisedProductName"].str.lower()

# Merge with budget tracker
merged_df = unmatched_df.merge(
    cleaned_budget_tracker,
    left_on="NormalisedProductName",
    right_on="Campaign",
    how="left"
)

merged_df = merged_df.drop(columns=['NormalisedProductName', 'Campaign_y'])
merged_df = merged_df.rename(columns={"Campaign_x": "Campaign"})

# Extract last budget value
last_budget_value = annual_budget["2024 CURRENT FORECAST Year 2025 Budget"].dropna().iloc[-1]
last_fnb_index = merged_df[merged_df["Division"] == "F&B"].index.max()

# Add new row for budget
new_row = pd.DataFrame({
    "Division": ["F&B"],
    "ChanelBudget": [last_budget_value]
})

merged_df = pd.concat([merged_df, new_row], ignore_index=True)
