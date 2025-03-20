import os
import time
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from azure.kusto.data import KustoClient, KustoConnectionStringBuilder, DataFormat
from azure.kusto.data.helpers import dataframe_from_result_table
from azure.kusto.ingest import QueuedIngestClient, IngestionProperties
from azure.kusto.ingest.ingestion_properties import ReportLevel
import tempfile

def save_dataframe_to_excel(df, filename, divisions):
    """
    Saves the DataFrame to an Excel file, splitting it into sheets by divisions.
    """
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for division in divisions:
            df_division = df[df['Division'] == division]
            sheet_name = str(division).replace(':', '').replace('\\', '').replace('/', '').replace('?', '').replace('*', '')[:31]
            if sheet_name.strip() == '':
                sheet_name = 'Other'
            df_division.to_excel(writer, sheet_name=sheet_name, index=False)
    time.sleep(2)  # Ensure the file is written


def delete_column_from_sheets(workbook, column_title):
    """
    Deletes a column from all sheets in the workbook based on the column title.
    """
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        column = None
        for cell in ws[1]:
            if cell.value == column_title:
                column = cell.column
                break
        if column:
            ws.delete_cols(column)

def format_workbook(workbook):
    """
    Applies formatting to all sheets in the workbook.
    """
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for col in ws.columns:
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True, size=12)
                
def set_column_width(sheet, start_col, headers, months):
    """
    Set the width of the columns based on the header text.
    """
    num_columns = len(headers)
    for i, month in enumerate(months):
        month_start_col = start_col + i * num_columns
        for j in range(num_columns):
            col_letter = get_column_letter(month_start_col + j)
            sheet.column_dimensions[col_letter].width = max(len(header) for header in headers) + 2  # Adding extra space for padding

def fill_colours(filename):
    wb = load_workbook(filename)
    delete_column_from_sheets(wb, "Division")
    # Define fill colours
    closedown_fill = PatternFill(start_color="E0DCDC", end_color="E0DCDC", fill_type="solid") # POCloseDownDate fill
    remainingpo_fill = PatternFill(start_color="E0ECF4", end_color="E0ECF4", fill_type="solid") # POValueRemaining fill
    date_fill = PatternFill(start_color="F8F4F4", end_color="F8F4F4", fill_type="solid") # date columns fill
    channel_fill = PatternFill(start_color="71AD47", end_color="71AD47", fill_type="solid")  # Green for Channel column
    product_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Light pink for product-related columns
    invoice_fill = PatternFill(start_color="C8E4B4", end_color="C8E4B4", fill_type="solid")  # Light green for TotalInvoiceVal column
    total_row_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")  # Gray for total rows
    header_fill = PatternFill(start_color="8095BF", end_color="8095BF", fill_type="solid")  # Header background
    header_font = Font(bold=True)  # White header text

    
    # Columns to format
    product_columns = ["ProductCode", "PlannedSpend", "ReservedBudget", "TotalBudget", "NetBillable", "AgencyCommission", "LevyASBOF", "TotalPOValue", "InvoiceNo"]
    date_columns = ["StartDate", "EndDate"]
    invoice_columns = ["TotalInvoicedToDate","TotalInvoiceVal"]
    
    # Monthly summary sheet names
    monthly_sheets = ["F&B Monthly", "FSH&EW Monthly", "W&FJ Monthly"]
    
    # Apply formatting to all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Apply styles to the header row (assumes headers are in row 1)
        for cell in ws[1]:  # Iterate over header row
            cell.fill = header_fill
            cell.font = header_font
        
        # Get column indices dynamically
        # column_indices = {cell.value: cell.column for cell in ws[1] if cell.value}
        
        # Apply formatting to columns if they exist
        
        if(sheet_name in monthly_sheets):
            header_row = 2
        else:
            header_row = 1
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            total_row = False  # Flag to check if row contains 'Total'
            
            for cell in row:
                col_name = ws.cell(row=header_row, column=cell.column).value  # Get column name
                cell_value = str(cell.value).lower() if cell.value else ""  # Convert value to lowercase for comparison
                    
                if col_name == "Channel":  # Channel column (green)
                    cell.fill = channel_fill
                    
                elif col_name in product_columns:  # Product-related columns (light pink)
                    cell.fill = product_fill
                
                elif col_name in invoice_columns:  # TotalInvoiceVal column (light green)
                    cell.fill = invoice_fill
                
                elif col_name == "POValueRemaining":
                    cell.fill = remainingpo_fill
                    
                elif col_name == "POCloseDownDate":
                    cell.fill = closedown_fill
                    
                elif col_name in date_columns:
                    cell.fill = date_fill
                    
                if "total" in cell_value:  # Check if 'Total' appears in any column of the row
                    total_row = True
                    
                # If the row is a total row, apply gray to all its cells
                if total_row:
                    for cell in row:
                        cell.fill = total_row_fill
                        
    format_workbook(wb)
    wb.save(filename)
    
def merge_duplicates(wb):
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Define column indices
        column_indices = {
            "PO_Number": 1,  
            "StartDate": 2,  
            "EndDate": 3,  
            "POCloseDownDate": 4, 
            "Campaign": 6,
            "Channel": 7
            }
        
        columns_to_merge_by_value = {"PO_Number", "Campaign", "Channel"}  # Columns that should be merged based on values
        columns_to_merge_by_structure = {"StartDate", "EndDate", "POCloseDownDate"}  # Empty columns to follow PO_number merge structure
        
        # Track merge ranges based on PO_number
        po_values = [ws.cell(row=row, column=column_indices["PO_Number"]).value for row in range(2, ws.max_row + 1)]
        merge_ranges = []  # Stores (start_row, end_row)
        
        start_row = 2  # Data starts from row 2
        for i in range(1, len(po_values)):
            if po_values[i] != po_values[i - 1]:  # If PO_number changes, merge previous block
                if i + 1 > start_row:  # Ensure more than 1 row exists
                    merge_ranges.append((start_row, i + 1))  # Store the range
                start_row = i + 2  # Move to next block
                
        # Merge last block if needed
        if start_row <= ws.max_row:
            merge_ranges.append((start_row, ws.max_row))
            
        # **1. Merge columns based on values**
        for col_name in columns_to_merge_by_value:
            col_index = column_indices[col_name]
            values = [ws.cell(row=row, column=col_index).value for row in range(2, ws.max_row + 1)]
            
            start_row = 2
            for i in range(1, len(values)):
                if values[i] != values[i - 1]:  
                    if i + 1 > start_row:
                        ws.merge_cells(start_row=start_row, end_row=i + 1, start_column=col_index, end_column=col_index)
                    start_row = i + 2  
                    
            if start_row <= ws.max_row:
                ws.merge_cells(start_row=start_row, end_row=ws.max_row, start_column=col_index, end_column=col_index)
                
        # **2. Apply PO_number merge structure to empty columns**
        for col_name in columns_to_merge_by_structure:
            col_index = column_indices[col_name]
            for start_row, end_row in merge_ranges:
                ws.merge_cells(start_row=start_row, end_row=end_row, start_column=col_index, end_column=col_index)
    
    

# Define ADX information
CLUSTER = "https://chanelmediacluster.uksouth.kusto.windows.net"
DATABASE = "financial-database-1"
BILLED_QUERY = """
//further clean billed report

billed_report
// add column of months
| extend Month = datetime_part("month", todatetime(BuyMonth))
| extend Month = case(
    Month == 1, "January",
    Month == 2, "February",
    Month == 3, "March",
    Month == 4, "April",
    Month == 5, "May",
    Month == 6, "June",
    Month == 7, "July",
    Month == 8, "August",
    Month == 9, "September",
    Month == 10, "October",
    Month == 11, "November",
    Month == 12, "December",
    "Unknown"
)
// add column of media names from PO tracker
| extend Channel = case(
    ClientCode == "C60" or ClientCode == "C65", "PPC",
    MediaName == "SEARCH&SOC" and ClientCode != "C60" and ClientCode != "C65", "Social",
    MediaName == "DISPLAY" and CampaignName contains "VIDEO", "Video",
    MediaName == "DISPLAY" and CampaignName !contains "VIDEO", "Display",
    MediaName == "PRESS", "Print",
    MediaName == "CINEMA", "Cinema",
    MediaName == "TELEVISION", "TV",
    MediaName == "FEES", "Retainer Fee",
    MediaName == "POSTER", "OOH",
    MediaName == "INT'L", "Retainer Fee",
    MediaName)
// normalise product names
| extend NormalisedProductName = case(
    ProductName == "EYEWEAR", "Eyewear",
    ProductName contains "FASHION" or ProductName contains "MDA", "Fashion",
    ProductName == "JEWELLERY HJ", "High Jewellery",
    ProductName contains "WATCHES", "Watches",
    ProductName == "BLEU", "Bleu",
    ProductName == "UK", "PPC",
    ProductName == "IRE" or ProductName == "ROI", "PPC",
    ProductName contains "MAKE UP" or ProductName contains "LES BEIGE", "Make Up",
    ProductName contains "SKINCARE", "Skincare",
    ProductName contains "CHANCE", "Chance",
    ProductName contains "COCO MELLE", "Coco Melle",
    ProductName contains "NO 5", "No 5",
    ProductName)
// create market column based on LocalCurrencyCode
| extend Market = case(
    ClientCode == "C50" or ClientCode == "C51" or ClientCode == "C52" or ClientCode == "C60", "UK",
    ClientCode == "C58" or ClientCode == "C65", "IRE",
    "")
// clean up campaign name - remove _ and CHANEL, remove channel name
| extend Campaign = replace_regex(CampaignName, @"_", " ")
| extend Campaign = replace_regex(Campaign, @"\b(CHANEL)\b", "")
| extend Campaign = replace_regex(Campaign, @"-", " ")
| extend Campaign = replace_regex(Campaign, @"\b(UK|IRE|SOCIAL|SOCIA|SO|SOC|DISPLAY|D|DSP|DISP|FEE|FEES)\b", "")
// create year column, and create MainName column without year and suffix
| extend Year = extract(@"(20\d{2}|2[0-9])\b", 0, Campaign)
| extend Year = iff(strlen(Year) == 2, strcat("20", Year), Year)
// remove unnecessary characters
| extend Campaign = trim(" ", replace_regex(Campaign, @"(20\d{2}|2[0-9])\b|FB|FR|MU|WFJ|\bWA\b|FSH|BEA|VIDEO|\.", ""))
// remove spaces
| extend Campaign = trim(" ", replace_regex(Campaign, @" ", ""))
// normalise campaign names
| extend Campaign = case(
    Campaign contains "BLEUH1", "Bleu H1",
    Campaign contains "BLEUH2", "Bleu H2",
    Campaign contains "LESBEIGEHERO" or Campaign contains "LESBEIGESHERO", "Les Beiges HERO",
    Campaign contains "COCOCRUSHH1", "Coco Crush H1",
    Campaign contains "J12H1", "J12 H1",
    Campaign contains "N1H2", "N1 H2",
    Campaign contains "EYEWEAR", "Eyewear",
    Campaign contains "NO5LEAU", "No.5 Leau",
    Campaign contains "C50/07", "C50/07",
    Campaign contains "NO5LEAUDROP", "No.5 Leau Drop",
    Campaign contains "BLEU", "Bleu",
    Campaign contains "LESBEIGES" or Campaign contains "LESBEIGE", "Les Beiges",
    Campaign contains "HIGHJEWELLERY", "High Jewellery",
    Campaign contains "COCOMELLE", "Coco Melle",
    ProductName == "MDA", "MDA",
    Campaign)
|extend Division = case(
    Channel == "Retainer Fee" and Campaign contains "F&B", "F&B",
    Channel == "Retainer Fee" and Campaign contains "FASHION", "FSH&EW",
    Channel == "Retainer Fee" and Campaign contains "EYEWEAR", "FSH&EW",
    Channel == "Retainer Fee" and Campaign contains "JEWELLERY", "W&FJ",
    Channel == "Retainer Fee" and Campaign contains "WATCHES", "W&FJ",
    Division == "Paid Search", "F&B",
    Channel == "Retainer Fee", Division,  // Keep the original division if no match
    Division  // Default to the original division if Channel is not "Retainer Fee"
)
| project 
    PO_Number = PO,
    Campaign,
    Channel,
    ProductCode,
    NormalisedProductName,
    NetBillable = Payable,
    AgencyCommission,
    LevyASBOF = LevyBillable,
    Total_Invoice_Val = UnbilledClientCost,
    Market,
    Division,
    Month,
    InvoiceNo
| sort by Division asc, Campaign asc
"""
BUDGET_QUERY = """
budgettracker
| extend GrandTotalReserve = ['GRAND TOTAL inc reserve'],
PlannedSpendData = ['Planned Spend latest plan'],
Reserve = ['Reserve'] // Include the reserve amount
| extend TotalAvailableBudget = GrandTotalReserve // Budget already includes reserve
| extend Campaign = tolower(trim(" ", Campaign))
| join kind=leftouter (
billed_report
| extend Campaign = replace_regex(CampaignName, @"_", " ")
| extend Campaign = replace_regex(Campaign, @"\b(CHANEL)\b", "")
| extend Campaign = replace_regex(Campaign, @"-", " ")
| extend Campaign = replace_regex(Campaign, @"\b(UK|IRE|SOCIAL|DISPLAY|FEE|DSP|DISP)\b", "")
| extend Campaign = trim(" ", replace_regex(Campaign, @"(20\d{2})\b|FB|FR|MU|WFJ|WA|FSH|BEA|VIDEO|\.", ""))
| extend Campaign = trim(" ", replace_regex(Campaign, @"\s+", " "))
| extend Campaign = tolower(Campaign)
| extend Campaign = case(
Campaign contains "bleu", "Bleu",
Campaign contains "no.5", "No. 5",
Campaign contains "les beige", "Make Up",
Campaign contains "coco melle", "Coco Melle",
Campaign contains "chance", "Chance",
Campaign contains "skincare", "Skincare",
Campaign contains "travel", "Travel Retail",
Campaign contains "watches", "Watches",
Campaign contains "jewellery", "Fine Jewellery",
Campaign contains "high jewellery", "High Jewellery",
Campaign contains "fashion", "Fashion",
Campaign contains "eyewear", "Eyewear",
Campaign contains "ppc", "PPC",
"Other")
| project Campaign, UnbilledCurrency
) on Campaign
| summarize
TotalUnbilled = sum(UnbilledCurrency),
TotalBudget = sum(TotalAvailableBudget),
PlannedSpend = sum(PlannedSpendData),
ReservedBudget = sum(Reserve)
by Campaign, Market
| project Campaign, PlannedSpend, ReservedBudget, TotalBudget, Market
"""

ANNUAL_BUDGET_QUERY = "annual_budget_sheet"

# Authentication (Choose the appropriate method)
KCSB = KustoConnectionStringBuilder.with_az_cli_authentication(CLUSTER)
KCSB_INGEST = KustoConnectionStringBuilder.with_az_cli_authentication(CLUSTER.replace("https://", "https://ingest-"))

# Create Kusto Client
client = KustoClient(KCSB)

def ingest_summary(KCSB_INGEST, summary):
    
    ingest_client = QueuedIngestClient(KCSB_INGEST)
    
    # Save DataFrame as CSV for ingestion
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as temp_file:
        summary.to_csv(temp_file.name, header=None, index=False)  # Save DataFrame to CSV
        temp_filename = temp_file.name
        
        # Define Ingestion Properties
        ingestion_props = IngestionProperties(
            database=DATABASE,
            table="new_summary",
            data_format=DataFormat.CSV,  # You can also use JSON or Parquet
            report_level=ReportLevel.FailuresAndSuccesses,
            flush_immediately=True  # Ensure data is available quickly
            )
        
        # Ingest the CSV file into ADX
        ingest_client.ingest_from_file(temp_filename, ingestion_properties=ingestion_props)


# Execute the query
response_1 = client.execute(DATABASE, BILLED_QUERY)
response_2 = client.execute(DATABASE, BUDGET_QUERY)
response_3 = client.execute(DATABASE, ANNUAL_BUDGET_QUERY)

data = dataframe_from_result_table(response_1.primary_results[0])

# Convert response to Pandas DataFrame
cleaned_billed = dataframe_from_result_table(response_1.primary_results[0])
cleaned_budget_tracker = dataframe_from_result_table(response_2.primary_results[0])
annual_budget = dataframe_from_result_table(response_3.primary_results[0])

pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", None) 

# print(cleaned_budget_tracker)
# print(cleaned_billed)

# Append total rows for each campaign
overall_totals = cleaned_billed.groupby(['PO_Number', 'Market','Division']).agg({
    'Campaign': 'first',
    'NetBillable': 'sum',
    'AgencyCommission': 'sum',
    'LevyASBOF': 'sum',
    'Total_Invoice_Val': 'sum',
    'ProductCode': 'first',
    'NormalisedProductName': 'first',
}).reset_index()
overall_totals['Channel'] = "Total"

# Append total rows for each campaign
monthly_totals = cleaned_billed.groupby(['PO_Number', 'Month', 'Market','Division']).agg({
    'Campaign': 'first',
    'NetBillable': 'sum',
    'AgencyCommission': 'sum',
    'LevyASBOF': 'sum',
    'Total_Invoice_Val': 'sum',
    'ProductCode': 'first',
    'NormalisedProductName': 'first',
}).reset_index()
monthly_totals['Channel'] = "Total"
 
channel_totals = cleaned_billed.groupby(['Campaign','Channel','Market','Division']).agg({
    'NetBillable': 'sum',
    'AgencyCommission': 'sum',
    'LevyASBOF': 'sum',
    'Total_Invoice_Val': 'sum',
    'PO_Number': 'first'
}).reset_index()
 
overall_totals['InvoiceNo'] = 'Total'
overall_totals['Month'] = 'Total'
 
# sort dataframes
overall_totals = overall_totals.sort_values(by=['Division', 'Campaign', 'Market'])
sorted_billed = cleaned_billed.sort_values(by=['Division', 'Campaign', 'Market'])
 
# make campaign names lowercase
overall_totals['NormalisedProductName'] = overall_totals['NormalisedProductName'].str.lower()
cleaned_billed['NormalisedProductName'] = cleaned_billed['NormalisedProductName'].str.lower()
 
# dataframe to use for monthly summary
monthly_summary = pd.concat([cleaned_billed, monthly_totals], ignore_index=True)
 
# Create a custom sort key that moves "Total" rows to the bottom
monthly_summary['is_total'] = monthly_summary['Channel'].str.contains('Total', na=False)

# Find PO number(s) for Retainer Fee
po_numbers = monthly_summary.loc[monthly_summary["Channel"] == "Retainer Fee", "PO_Number"].tolist()
monthly_summary["is_fee"] = monthly_summary["PO_Number"].isin(po_numbers)
 
# Sort by PO_Number, then by the 'is_total' flag (False first, then True), then by Campaign (if needed)
monthly_summary = monthly_summary.sort_values(by=['Division', 'is_fee', 'PO_Number', 'is_total', 'Campaign', 'Channel', 'Month'], ascending=[True, True, True, True, True, True, True]) 

# Drop columns
monthly_summary = monthly_summary.drop(columns=['is_total'])
monthly_summary = monthly_summary.drop(columns=['NormalisedProductName'])
monthly_summary = monthly_summary.drop(columns=['is_fee'])
 
 
# Merge totals_df with cleaned_budget_tracker based on NormalisedProductName and Campaign
merged_totals_df = overall_totals.merge(
    cleaned_budget_tracker,
    left_on=["NormalisedProductName","Market"],
    right_on=["Campaign","Market"],
    how="left"
)


# print(merged_totals_df)

non_totals_df = cleaned_billed.sort_values(by=['Division', 'Campaign', 'Channel', 'Market'])
 
# sort and rename values
merged_totals_df = merged_totals_df.sort_values(by=['Division', 'PO_Number', 'Campaign_x'])
merged_totals_df.rename(columns={"Campaign_x": "Campaign"}, inplace=True)
 
# create merged df with non total and total values
final_merged_df = pd.concat([non_totals_df, merged_totals_df], ignore_index=True)
 
# Create merged df with channel totals and overall totals
channel_summary = pd.concat([channel_totals, merged_totals_df], ignore_index=True)
 
# Create a custom sort key that moves "Total" rows to the bottom and "Fee" rows to the bottom
channel_summary['is_total'] = channel_summary['Channel'].str.contains('Total', na=False)

# Find PO number(s) for Retainer Fee
po_numbers = channel_summary.loc[channel_summary["Channel"] == "Retainer Fee", "PO_Number"].tolist()
channel_summary["is_fee"] = channel_summary["PO_Number"].isin(po_numbers)
 
# Sort 
channel_summary = channel_summary.sort_values(by=['Division', 'is_fee','PO_Number', 'is_total', 'Campaign', 'Channel'], ascending=[True, True, True, True, True, True])
 
# Drop temporary columns
channel_summary = channel_summary.drop(columns=['is_total'])
channel_summary = channel_summary.drop(columns=['is_fee'])
 
# Sort by Division, Campaign, and Channel (ensuring totals appear last)
final_merged_df = final_merged_df.sort_values(by=['Division', 'Campaign', 'Channel'])
 
# Reset index
final_merged_df = final_merged_df.reset_index(drop=True)
final_merged_df = final_merged_df.drop(columns=['NormalisedProductName', 'Campaign_y'])

# define columns to keep for summary_df and channel_summary 
columns_to_keep = ['PO_Number', 'Campaign', 'Channel', 'ProductCode', 'TotalBudget', 'NetBillable', 'AgencyCommission', 'LevyASBOF', 'Total_Invoice_Val', 'Market', 'Division', 'InvoiceNo', 'Month']
columns_to_keep_channel = ['PO_Number', 'Campaign', 'Channel', 'PlannedSpend', 'ReservedBudget', 'TotalBudget', 'NetBillable', 'AgencyCommission', 'LevyASBOF', 'Total_Invoice_Val', 'Market', 'Division', 'InvoiceNo']
 
 
# only keep required columns
summary_df = final_merged_df[columns_to_keep]
channel_summary = channel_summary[columns_to_keep_channel]
 
# Drop duplicates
summary_df = summary_df.drop_duplicates()
 
# Add date columns
monthly_summary.loc[:, "StartDate"] = None
monthly_summary.loc[:, "EndDate"] = None
monthly_summary.loc[:, "POCloseDownDate"] = None
channel_summary.loc[:, "StartDate"] = None
channel_summary.loc[:, "EndDate"] = None
channel_summary.loc[:, "POCloseDownDate"] = None

monthly_order = ["PO_Number", "StartDate", "EndDate", "POCloseDownDate", "Market", "Campaign", "Channel", "NetBillable", "AgencyCommission", "LevyASBOF", "Total_Invoice_Val", "InvoiceNo","Division", "Month"]

ordered_monthly_summary = monthly_summary[monthly_order]

# print(ordered_monthly_summary)

# Add TotalPOValue and POValueRemaining columns
summary_df['TotalPOValue'] = summary_df['AgencyCommission'] + summary_df['LevyASBOF'] + summary_df['TotalBudget']
summary_df['POValueRemaining'] = summary_df['TotalPOValue'] - summary_df['Total_Invoice_Val']
channel_summary['TotalPOValue'] = channel_summary['AgencyCommission'] + channel_summary['LevyASBOF'] + channel_summary['TotalBudget']
channel_summary['POValueRemaining'] = channel_summary['TotalPOValue'] - channel_summary['Total_Invoice_Val']
 
# Create a custom sort key that moves "Total" rows to the bottom
summary_df['is_total'] = summary_df['Channel'].str.contains('Total', na=False)
 
# Sort by PO_Number, then by the 'is_total' flag (False first, then True), then by Campaign (if needed)
summary_df = summary_df.sort_values(by=['Division', 'PO_Number', 'is_total', 'Campaign', 'Channel'], ascending=[True, True, True, True, True])
 
# Drop the temporary 'is_total' column
summary_df = summary_df.drop(columns=['is_total'])
 
pd.set_option('display.max_columns', None)
# print(summary_df)
 
# Ensure budget-related values only appear where Channel == "Total"
columns_to_clear = [
    "TotalPOValue", "Total_Invoice_Val", "POValueRemaining", "TotalBudget", "PlannedSpend", "ReservedBudget", "ChanelBudget"]
 
summary_df.loc[summary_df["Channel"] != "Total", columns_to_clear] = None  # Set non-"Total" rows to NaN

# Set Campaign to None in total rows 
channel_summary.loc[channel_summary["Channel"] == "Total", ["Campaign"]] = None 
monthly_summary.loc[monthly_summary["Channel"] == "Total", ["Campaign"]] = None 
 
# Identify campaigns that appear at least three times
campaign_counts = summary_df['Campaign'].value_counts()
repeated_campaigns = campaign_counts[campaign_counts >= 3].index
 
# Process only campaigns that need merging
summary_df['PO_Number'] = summary_df.apply(
    lambda row: row['PO_Number'] if (row['Campaign'] not in repeated_campaigns or pd.notna(row['PO_Number'])) else row['PO_Number'], axis=1
)
 
summary_df['Campaign'] = summary_df.apply(
    lambda row: row['Campaign'] if (row['Campaign'] not in repeated_campaigns or pd.notna(row['Campaign'])) else row['Campaign'], axis=1
)
 
# Set non-numeric values to None where 'InvoiceNo' is "Total"
non_numeric_cols = ['PO_Number', 'Campaign', 'Market', 'Division']
summary_df.loc[summary_df['Channel'] == "Total", non_numeric_cols] = None
 
ordered_columns = ['PO_Number', 'Campaign', 'Channel', 'ProductCode', 'TotalBudget', 'NetBillable', 'AgencyCommission', 'LevyASBOF', 'TotalPOValue', 'Total_Invoice_Val', 'POValueRemaining', 'PlannedSpend', 'ReservedBudget', 'Market', 'Division', 'InvoiceNo', 'Month']
ordered_columns_channel = ['PO_Number', 'StartDate', 'EndDate', 'POCloseDownDate', 'Market', 'Campaign', 'Channel', 'PlannedSpend', 'ReservedBudget', 'TotalBudget', 'NetBillable', 'AgencyCommission', 'LevyASBOF', 'TotalPOValue', 'Total_Invoice_Val', 'POValueRemaining', 'Division']
 
ordered_summary_df = summary_df[ordered_columns]
channel_summary = channel_summary[ordered_columns_channel]

# rename Total_Invoice_Val to TotalInvoicedToDate
channel_summary.rename(columns={"Total_Invoice_Val": "TotalInvoicedToDate"}, inplace=True)

# print(channel_summary)
 
ingest_summary_df = ordered_summary_df

# Save DataFrame to Excel, divided by 'Division'
filename = "FormattedAnnualBudget.xlsx"
divisions = channel_summary['Division'].unique()

save_dataframe_to_excel(channel_summary, filename, divisions)


# Usage
filename = 'FormattedAnnualBudget.xlsx'  # Specify the path to your file
start_col = 19  # Adjust as necessary to the column where the January headers should begin
# print(summary_df)
# append_monthly_tables_to_excel(filename, start_col)
ingest_summary(KCSB_INGEST, ingest_summary_df)

def save_dataframe_to_excel(df, filename, divisions):
    """
    Saves the DataFrame to an Excel file, splitting it into sheets by divisions.
    """
    with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
        for division in divisions:
            df_division = df[df['Division'] == division]
            sheet_name = str(division).replace(':', '').replace('\\', '').replace('/', '').replace('?', '').replace('*', '')[:31]
            if sheet_name.strip() == '':
                sheet_name = 'Other'
            df_division.to_excel(writer, sheet_name=sheet_name, index=False)

def create_monthly_summary_sheet(filename, data):
    """
    Creates the 'Monthly Summary' sheet with:
    - Merged super headers for each month.
    - 4 subheaders per month: NetBillable, AgencyCommission, LevyASBOF, TotalInvoicedToDate.
    - Each row correctly mapped to its respective values.
    - The total is placed at the end for each Campaign-Channel.
    """
    # Load existing workbook
    book = load_workbook(filename)

    # Define headers
    base_headers = ['PO_Number', 'StartDate', 'EndDate', 'POCloseDownDate', 'Market', 'Campaign', 'Channel']
    months = ['January', 'February', 'March', 'April', 'May', 'June', 
              'July', 'August', 'September', 'October', 'November', 'December']
    metric_headers = ['NetBillable', 'AgencyCommission', 'LevyASBOF', 'TotalInvoiceVal', 'InvoiceNo']
    
    # Iterate over each unique Division
    for division in data["Division"].unique():
        division_data = data[data["Division"] == division]  # Filter data for this division

        sheet_name = f"{division} Monthly"
        
        if sheet_name in book.sheetnames:
            del book[sheet_name]
            
        sheet = book.create_sheet(sheet_name)
        
        

        # Set up merged headers for months
        row1 = 1  # Super header row
        row2 = 2  # Subheader row

        col_idx = 1  # Start column
        for header in base_headers:
            sheet.cell(row=row2, column=col_idx, value=header).font = Font(bold=True)
            sheet.cell(row=row2, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

        # Merge and set month headers
        for month in months:
            start_col = col_idx
            end_col = col_idx + len(metric_headers) - 1
            sheet.merge_cells(start_row=row1, start_column=start_col, end_row=row1, end_column=end_col)
            month_cell = sheet.cell(row=row1, column=start_col, value=month)
            month_cell.font = Font(bold=True, size=12)
            month_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Set subheaders for each month
            for metric in metric_headers:
                sheet.cell(row=row2, column=col_idx, value=metric).font = Font(bold=True)
                sheet.cell(row=row2, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                col_idx += 1

        # **Populate Data Correctly (Fixing Repetitions)**
        row_idx = 3  # Start inserting data from the 3rd row
        campaign_channel_groups = data.groupby(['Campaign', 'Channel'])

        # Start inserting data from row 3
        row_idx = 3  

        # Iterate over each unique Campaign + Channel combination
        for (campaign, channel), group_data in division_data.groupby(['Campaign', 'Channel']):
            col_idx = 1

            row_idx = 3  # Start inserting data from the 3rd row

            for _, row in division_data.iterrows():
                col_idx = 1  # Reset column index for each row

                # ensure PO_Number, Campaign, and Channel are written for every row
                sheet.cell(row=row_idx, column=col_idx, value=row['PO_Number'])
                sheet.cell(row=row_idx, column=col_idx + 4, value=row['Market'])
                sheet.cell(row=row_idx, column=col_idx + 5, value=row['Campaign'])
                sheet.cell(row=row_idx, column=col_idx + 6, value=row['Channel'])

                col_idx += 7  # Move to the monthly data columns

                # Retrieve and correctly assign the row's data based on the Month
                net_billable = None if pd.isna(row['NetBillable']) else row['NetBillable']
                agency_commission = None if pd.isna(row['AgencyCommission']) else row['AgencyCommission']
                levy_asbof = None if pd.isna(row['LevyASBOF']) else row['LevyASBOF']
                total_invoiced = None if pd.isna(row['Total_Invoice_Val']) else row['Total_Invoice_Val']
                invoice_no = None if pd.isna(row['InvoiceNo']) else row['InvoiceNo']


                # Assign values to the correct month
                if row['Month'] in months:  # Ensure the Month is valid before indexing
                    month_col_offset = (months.index(row['Month']) * 5)  # Offset based on month position
                    sheet.cell(row=row_idx, column=col_idx + month_col_offset, value=net_billable)
                    sheet.cell(row=row_idx, column=col_idx + month_col_offset + 1, value=agency_commission)
                    sheet.cell(row=row_idx, column=col_idx + month_col_offset + 2, value=levy_asbof)
                    sheet.cell(row=row_idx, column=col_idx + month_col_offset + 3, value=total_invoiced)

                sheet.cell(row=row_idx, column=col_idx + month_col_offset, value=net_billable)
                sheet.cell(row=row_idx, column=col_idx + month_col_offset + 1, value=agency_commission)
                sheet.cell(row=row_idx, column=col_idx + month_col_offset + 2, value=levy_asbof)
                sheet.cell(row=row_idx, column=col_idx + month_col_offset + 3, value=total_invoiced)
                sheet.cell(row=row_idx, column=col_idx + month_col_offset + 4, value=invoice_no)

                row_idx += 1  # Move to the next row for correct alignment

            for month in months:
                month_col_offset = months.index(month) * 4  # Offset based on month position
                col_idx += 4  # Move to the next set of columns

            row_idx += 1  # Move to next total row


        # Auto-adjust column width
        for col_num, column_cells in enumerate(sheet.columns, 1):
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in column_cells)
            adjusted_width = max(max_length + 2, 12)
            sheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    # Save and close workbook
    book.save(filename)
    book.close()
    print("'Monthly Summary' sheet correctly filled with actual values, with totals at the end!")

# Assuming `final_merged_df` contains the relevant data
filename = "FormattedAnnualBudget.xlsx"
# print(summary_df)
create_monthly_summary_sheet(filename, monthly_summary)

fill_colours(filename)

if os.path.exists(filename) and os.path.getsize(filename) > 0:
    wb = load_workbook(filename)
    
    merge_duplicates(wb)
    
    wb.save(filename)
    print("Merged cells successfully.")

    # Assuming 'delete_column_from_sheets' and 'format_workbook' functions are defined as previous
    delete_column_from_sheets(wb, 'Division')
    format_workbook(wb)
    
else:
    print("Error: The file does not exist or is empty.")


# monthly_summary = monthly_summary.reset_index(inplace=False)
# print(monthly_summary)